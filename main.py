import streamlit as st
import streamlit.components.v1 as components
import pdfplumber
import pandas as pd
import io
import html
import re
import numpy_financial as npf
from openpyxl.utils import get_column_letter


USD_PREFIXES = ("INDON", "INDOIS")


def normalize_column_name(col, idx):
    if pd.isna(col):
        return f"_unnamed_{idx}"

    name = str(col).replace("\n", "").replace("\r", "").strip()
    return name or f"_unnamed_{idx}"


def normalize_columns(columns):
    normalized = []
    seen = {}

    for idx, col in enumerate(columns, start=1):
        name = normalize_column_name(col, idx)
        seen[name] = seen.get(name, 0) + 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        normalized.append(name)

    return normalized


def drop_repeated_header_rows(df):
    normalized_headers = [str(c).strip().lower() for c in df.columns]

    def looks_like_header(row):
        row_values = [str(v).replace("\n", "").replace("\r", "").strip().lower() for v in row.tolist()]
        return row_values == normalized_headers

    return df[~df.apply(looks_like_header, axis=1)].reset_index(drop=True)


def clean_numeric(val):
    if pd.isna(val) or val == "":
        return val

    if not isinstance(val, str):
        try:
            return float(val)
        except (TypeError, ValueError):
            return val

    s = re.sub(r"\s+", "", val.strip())
    if s in ("", "-"):
        return val

    is_percent = "%" in s
    s = s.replace("%", "")

    if not all(c in "0123456789.,-" for c in s):
        return val

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        if s.count(",") > 1:
            s = s.replace(",", "")
        else:
            parts = s.split(",")
            if len(parts[-1]) == 3:
                s = s.replace(",", "")
            else:
                s = s.replace(",", ".")

    try:
        v = float(s)
        return v / 100.0 if is_percent else v
    except ValueError:
        return val


def classify_currency(product_code):
    if not isinstance(product_code, str):
        return "IDR"

    code = product_code.strip().upper()
    return "US" if code.startswith(USD_PREFIXES) else "IDR"


def to_percent_str(val):
    try:
        if pd.isna(val):
            return None
        return f"{round(float(val), 6)}%"
    except (TypeError, ValueError):
        return val


def parse_percent_rate(val):
    try:
        if pd.isna(val):
            return None
        return float(str(val).replace("%", "").replace(",", ".")) / 100.0
    except (TypeError, ValueError):
        return None


def excel_ref(df, column_name, row_number):
    col_idx = df.columns.get_loc(column_name) + 1
    return f"{get_column_letter(col_idx)}{row_number}"

# Konfigurasi agar halaman tampil full-width (lebar penuh)
st.set_page_config(layout="wide", page_title="PDF Table Extractor & Editor")

st.title("PDF Table Extractor & Editor")

uploaded_file = st.file_uploader("Unggah file PDF Excel", type="pdf")

if uploaded_file is not None:
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            all_rows = []
            
            # Ekstrak tabel dari semua halaman
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    all_rows.extend(table)
            
            if all_rows:
                # Jadikan baris pertama sebagai header
                header = all_rows[0]
                data = all_rows[1:]
                
                df = pd.DataFrame(data, columns=header)
                
                # Bersihkan nama kolom dari karakter 'enter' dan beri nama aman untuk kolom kosong.
                df.columns = normalize_columns(df.columns)
                df = drop_repeated_header_rows(df)
                
                # Tambah kolom kustom di sebelah kanan tabel
                
                # --- PEMBERSIHAN DATA DAN KONVERSI TIPE ---
                # Konversi semua kolom numerik utama agar jadi float asli
                numeric_candidates = ['kupon', 'mbi_beli', 'mbi_jual', 'yield_mbi_beli', 'yield_mbi_jual', 'Inventory']
                for c in numeric_candidates:
                    if c in df.columns:
                        df[c] = df[c].apply(clean_numeric)
                
                # Konversi kolom tanggal ke Datetime
                date_candidates = ['Maturity', 'Settlement date']
                for c in date_candidates:
                    if c in df.columns:
                        df[c] = pd.to_datetime(df[c], errors='coerce')

                # 1. Currency Check
                if 'product_code' in df.columns:
                    df['currency check'] = df['product_code'].apply(classify_currency)
                
                # 2. Year Maturity
                maturity_col = next((col for col in df.columns if pd.notna(col) and 'maturity' in str(col).lower()), None)
                if maturity_col:
                    def extract_year(date_val):
                        if pd.isna(date_val): return ''
                        if isinstance(date_val, pd.Timestamp): return str(date_val.year)
                        match = re.search(r'((?:19|20)\d{2})', str(date_val))
                        return match.group(1) if match else str(date_val)
                    df['year maturity'] = df[maturity_col].apply(extract_year)
                
                # 3, 4, 5. Kolom Persentase (Kupon %, Y MBI Beli, Y MBI Jual)
                if 'kupon' in df.columns:
                    df['kupon %'] = df['kupon'].apply(to_percent_str)
                else:
                    df['kupon %'] = None
                
                if 'yield_mbi_beli' in df.columns:
                    df['y mbi beli'] = df['yield_mbi_beli'].apply(to_percent_str)
                
                if 'yield_mbi_jual' in df.columns:
                    df['y mbi jual'] = df['yield_mbi_jual'].apply(to_percent_str)
                else:
                    df['y mbi jual'] = None
                    st.warning("Kolom yield_mbi_jual tidak ditemukan; simulasi yield dan grafik akan kosong.")
                
                # 6. Kolom MDURATION 
                st.write("---")
                st.subheader("Pengaturan Parameter Simulasi")
                
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    settlement_date = st.date_input("Tanggal Settlement ($Y$4)", value=pd.to_datetime("today"))
                with col2:
                    cut_rate_input = st.number_input("Cut / Hike Rate (%)", value=0.2500, step=0.0100, format="%.4f")
                with col3:
                    base_date_input = st.date_input("Tanggal Basis (Hari Ini)", value=pd.to_datetime("today"))
                with col4:
                    yield_hike_input = st.number_input("Yield Naik (%)", value=0.0, step=0.0100, format="%.4f")
                with col5:
                    yield_cut_input = st.number_input("Yield Turun (%)", value=0.0, step=0.0100, format="%.4f")
                
                def hitung_mduration_ql(row):
                    try:
                        import QuantLib as ql
                        
                        mat_val = row.get('Maturity') if 'Maturity' in row else (row.get(maturity_col) if maturity_col else None)
                        kup_val = row.get('kupon %')
                        yld_val = row.get('y mbi jual')
                        
                        if pd.isna(mat_val) or pd.isna(kup_val) or pd.isna(yld_val):
                            return None
                        
                        maturity_date = pd.to_datetime(mat_val)
                        settlement_dt = pd.to_datetime(settlement_date)
                        
                        if pd.isna(maturity_date) or maturity_date <= settlement_dt:
                            return None
                            
                        c = parse_percent_rate(kup_val)
                        y = parse_percent_rate(yld_val)
                        if c is None or y is None:
                            return None
                        
                        set_date_ql = ql.Date(settlement_dt.day, settlement_dt.month, settlement_dt.year)
                        mat_date_ql = ql.Date(maturity_date.day, maturity_date.month, maturity_date.year)
                        
                        ql.Settings.instance().evaluationDate = set_date_ql
                        
                        calendar = ql.NullCalendar()
                        day_count = ql.ActualActual(ql.ActualActual.ISMA) 
                        q_freq = ql.Semiannual
                        
                        schedule = ql.Schedule(set_date_ql, mat_date_ql, ql.Period(q_freq), calendar,
                                               ql.Unadjusted, ql.Unadjusted, ql.DateGeneration.Backward, False)
                        
                        bond = ql.FixedRateBond(0, 100.0, schedule, [c], day_count)
                        
                        interest_rate = ql.InterestRate(y, day_count, ql.Compounded, q_freq)
                        mod_dur = ql.BondFunctions.duration(bond, interest_rate, ql.Duration.Modified)
                        
                        return round(mod_dur, 4)
                    except Exception:
                        return None

                df['MDURATION'] = df.apply(hitung_mduration_ql, axis=1)

                # 7 & 8. Kolom Rate Hike & Rate Cut
                def hitung_rate_impact(val, is_hike=False):
                    try:
                        if pd.isna(val): return None
                        v = parse_percent_rate(val)
                        if v is None:
                            return None
                        
                        result = v * (cut_rate_input / 100.0) 
                        if is_hike:
                            result = -result 
                            
                        return round(result, 6)
                    except Exception:
                        return None
                        
                df['Rate Hike'] = df['y mbi jual'].apply(lambda x: hitung_rate_impact(x, is_hike=True))
                df['Rate Cut'] = df['y mbi jual'].apply(lambda x: hitung_rate_impact(x, is_hike=False))

                if 'mbi_jual' in df.columns:
                    mbi_jual_num = pd.to_numeric(df['mbi_jual'], errors='coerce')
                    df['Rate Hike Price'] = mbi_jual_num + (df['Rate Hike'] * mbi_jual_num)
                    df['Rate Cut Price'] = mbi_jual_num + (df['Rate Cut'] * mbi_jual_num)
                else:
                    df['Rate Hike Price'] = None
                    df['Rate Cut Price'] = None
                
                base_date_pd = pd.to_datetime(base_date_input)
                
                # Kolom Total Year to Maturity
                def hitung_ytm_years(row):
                    mat = row.get('Maturity') if 'Maturity' in row else row.get(maturity_col)
                    if pd.isna(mat): return None
                    try:
                        diff = (pd.to_datetime(mat) - base_date_pd.normalize()).days / 365.25
                        return round(diff, 4)
                    except:
                        return None
                        
                df['Total Year to Maturity'] = df.apply(hitung_ytm_years, axis=1)
                
                # Kolom Price menggunakan formula PV
                def hitung_price_pv(row, is_hike=True):
                    try:
                        y_mbi = parse_percent_rate(row.get('y mbi jual'))
                        kupon = parse_percent_rate(row.get('kupon %'))
                        ytm_years = row.get('Total Year to Maturity')
                        
                        if pd.isna(y_mbi) or pd.isna(kupon) or pd.isna(ytm_years):
                            return None
                        
                        if is_hike:
                            rate = y_mbi + (yield_hike_input / 100.0)
                        else:
                            rate = y_mbi - (yield_cut_input / 100.0)
                            
                        nper = ytm_years
                        pmt = 100 * kupon
                        fv = 100
                        
                        price = -npf.pv(rate, nper, pmt, fv, when=0)
                        return round(price, 4)
                    except Exception:
                        return None
                        
                df['Price if Yield Hike'] = df.apply(lambda r: hitung_price_pv(r, is_hike=True), axis=1)
                df['Price if Yield Cut'] = df.apply(lambda r: hitung_price_pv(r, is_hike=False), axis=1)

                st.write("Edit data langsung pada tabel di bawah:")
                
                # Memastikan tabel memanfaatkan seluruh lebar kontainer
                edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True)

                # --- TAMPILAN WARNA TABEL ---
                st.write("Tabel dengan gradasi warna:")
                year_col = next((c for c in edited_df.columns if "year" in str(c).lower()), None)

                styled_df = edited_df.copy()
                styler = styled_df.style

                if year_col:
                    styler = styler.background_gradient(
                        subset=[year_col],
                        cmap="Blues",
                        gmap=pd.to_numeric(styled_df[year_col], errors="coerce")
                    )

                # Gelapkan sel numerik lain dengan gradasi abu-abu agar kontras lebih tinggi.
                numeric_cols = [
                    c for c in styled_df.columns
                    if c != year_col and pd.api.types.is_numeric_dtype(styled_df[c])
                ]
                if numeric_cols:
                    styler = styler.background_gradient(
                        subset=numeric_cols,
                        cmap="Greys"
                    )

                st.dataframe(styler, use_container_width=True, hide_index=True)

                # --- COPY TABLE (PRODUCT_CODE SAMPAI INVENTORY) ---
                copy_cols = []
                if "product_code" in edited_df.columns and "Inventory" in edited_df.columns:
                    start_idx = edited_df.columns.get_loc("product_code")
                    end_idx = edited_df.columns.get_loc("Inventory")
                    if start_idx <= end_idx:
                        copy_cols = edited_df.columns[start_idx:end_idx + 1].tolist()

                if copy_cols:
                    copy_df = edited_df[copy_cols].copy()
                    copy_text = copy_df.to_csv(index=False, sep="\t")
                    escaped_copy_text = html.escape(copy_text)
                    st.write("Copy tabel (kolom product_code sampai Inventory):")
                    components.html(
                        f"""
                        <div style='display:flex; gap:8px; align-items:center;'>
                          <button onclick='navigator.clipboard.writeText(document.getElementById("copy_area").value)'>Copy Table</button>
                          <span id='copy_status' style='font-size:12px; color:#333;'></span>
                        </div>
                        <textarea id='copy_area' style='width:100%; height:160px; margin-top:8px;'>{escaped_copy_text}</textarea>
                        <script>
                          const btn = document.querySelector('button');
                          const status = document.getElementById('copy_status');
                          btn.addEventListener('click', () => {{
                            status.textContent = 'Copied';
                            setTimeout(() => status.textContent = '', 1500);
                          }});
                        </script>
                        """,
                        height=230,
                    )
                else:
                    st.info("Kolom product_code atau Inventory tidak ditemukan untuk fitur copy.")

                excel_df = edited_df.copy()

                # --- VISUALISASI YIELD CURVE BENCHMARK ---
                st.write("---")
                
                if 'product_code' in edited_df.columns and 'year maturity' in edited_df.columns and 'y mbi jual' in edited_df.columns:
                    
                    if 'currency check' in edited_df.columns:
                        available_currencies = ["Semua"] + edited_df['currency check'].dropna().unique().tolist()
                        selected_currency = st.radio("Filter Mata Uang Grafik:", options=available_currencies, horizontal=True)
                    else:
                        selected_currency = "Semua"

                    chart_title = f"Bonds Chart {selected_currency if selected_currency != 'Semua' else ''} - Mark to Market".replace("  ", " ")
                    st.subheader(chart_title)
                    
                    filtered_df_for_chart = edited_df.copy()
                    if selected_currency != "Semua" and 'currency check' in filtered_df_for_chart.columns:
                        filtered_df_for_chart = filtered_df_for_chart[filtered_df_for_chart['currency check'] == selected_currency]
                        
                    valid_products = filtered_df_for_chart['product_code'].dropna().unique().tolist()
                    
                    bench_col1, bench_col2 = st.columns(2)
                    with bench_col1:
                        benchmark_selection_1 = st.multiselect(
                            "Pilih Seri Obligasi Benchmark 1",
                            options=valid_products,
                            default=[],
                            key="benchmark_selection_1",
                        )
                    with bench_col2:
                        benchmark_selection_2 = st.multiselect(
                            "Pilih Seri Obligasi Benchmark 2",
                            options=valid_products,
                            default=[],
                            key="benchmark_selection_2",
                        )
                    
                    chart_df = filtered_df_for_chart.copy()
                    chart_df['Year Numeric'] = pd.to_numeric(chart_df['year maturity'], errors='coerce')
                    
                    def parse_yield_chart(val):
                        try:
                            return float(str(val).replace('%', '').replace(',', '.'))
                        except:
                            return None
                    chart_df['Yield Numeric'] = chart_df['y mbi jual'].apply(parse_yield_chart)
                    
                    chart_df = chart_df.dropna(subset=['Year Numeric', 'Yield Numeric'])
                    
                    if not chart_df.empty:
                        import plotly.graph_objects as go
                        
                        fig = go.Figure()
                        
                        # 1. Scatter Plot
                        fig.add_trace(go.Scatter(
                            x=chart_df['Year Numeric'],
                            y=chart_df['Yield Numeric'],
                            mode='markers+text',
                            name='Seri Bonds',
                            text=chart_df['product_code'],
                            textposition="top right",
                            marker=dict(size=8, color='#5D9CEC'),
                            textfont=dict(color='black', size=10) 
                        ))
                        
                        # 2. Line Plot (Benchmark Curves)
                        benchmark_configs = [
                            ("Benchmark 1", benchmark_selection_1, "red"),
                            ("Benchmark 2", benchmark_selection_2, "green"),
                        ]

                        summary_tables = []
                        for bench_name, selected_products, color in benchmark_configs:
                            if not selected_products:
                                continue

                            bench_df = chart_df[chart_df['product_code'].isin(selected_products)].copy()
                            bench_df = bench_df.sort_values(by='Year Numeric')

                            fig.add_trace(go.Scatter(
                                x=bench_df['Year Numeric'],
                                y=bench_df['Yield Numeric'],
                                mode='lines+markers',
                                name=f'{bench_name} Curve',
                                line=dict(color=color, width=3),
                                marker=dict(size=10, color=color)
                            ))

                            summary_df = bench_df[['product_code', 'year maturity', 'y mbi jual']].copy()
                            summary_df.columns = [bench_name, 'Year', 'Yield']
                            summary_tables.append((bench_name, summary_df))

                        if summary_tables:
                            st.write("**Tabel Ringkasan Benchmark**")
                            summary_cols = st.columns(len(summary_tables))
                            for idx, (bench_name, summary_df) in enumerate(summary_tables):
                                with summary_cols[idx]:
                                    st.write(f"**{bench_name}**")
                                    st.dataframe(summary_df, hide_index=True, use_container_width=True)
                            
                        # Layout Diperbaiki agar rapi saat ditarik full layar
                        fig.update_layout(
                            xaxis_title="Year",
                            yaxis_title="Yield (% MBI Jual)",
                            hovermode="closest",
                            yaxis=dict(tickformat=".2f", ticksuffix="%"),
                            height=600,
                            plot_bgcolor='white',
                            showlegend=True,
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1), # Legenda diletakkan di atas
                            margin=dict(l=20, r=20, t=40, b=20)
                        )
                        
                        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
                        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
                        
                        st.plotly_chart(fig, use_container_width=True)
                
                # --- PERSIAPAN DOWNLOAD EXCEL ---
                def clean_for_excel(val):
                    parsed = parse_percent_rate(val)
                    return parsed if parsed is not None else val
                        
                for c in ['kupon %', 'y mbi beli', 'y mbi jual']:
                    if c in excel_df.columns:
                        excel_df[c] = excel_df[c].apply(clean_for_excel)
                
                if {'Maturity', 'kupon %', 'y mbi jual'}.issubset(excel_df.columns):
                    excel_df['MDURATION'] = [
                        f"=MDURATION('_Parameters'!$B$1,{excel_ref(excel_df, 'Maturity', i+2)},{excel_ref(excel_df, 'kupon %', i+2)},{excel_ref(excel_df, 'y mbi jual', i+2)},2,1)"
                        for i in range(len(excel_df))
                    ]

                if 'y mbi jual' in excel_df.columns:
                    excel_df['Rate Hike'] = [
                        f"=-{excel_ref(excel_df, 'y mbi jual', i+2)}*{cut_rate_input/100.0}"
                        for i in range(len(excel_df))
                    ]
                    excel_df['Rate Cut'] = [
                        f"={excel_ref(excel_df, 'y mbi jual', i+2)}*{cut_rate_input/100.0}"
                        for i in range(len(excel_df))
                    ]

                if {'mbi_jual', 'Rate Hike', 'Rate Cut'}.issubset(excel_df.columns):
                    excel_df['Rate Hike Price'] = [
                        f"={excel_ref(excel_df, 'mbi_jual', i+2)}+({excel_ref(excel_df, 'Rate Hike', i+2)}*{excel_ref(excel_df, 'mbi_jual', i+2)})"
                        for i in range(len(excel_df))
                    ]
                    excel_df['Rate Cut Price'] = [
                        f"={excel_ref(excel_df, 'mbi_jual', i+2)}+({excel_ref(excel_df, 'Rate Cut', i+2)}*{excel_ref(excel_df, 'mbi_jual', i+2)})"
                        for i in range(len(excel_df))
                    ]
                
                if {'y mbi jual', 'Total Year to Maturity', 'kupon %'}.issubset(excel_df.columns):
                    excel_df['Price if Yield Hike'] = [
                        f"=-PV(({excel_ref(excel_df, 'y mbi jual', i+2)}+{yield_hike_input/100.0}),{excel_ref(excel_df, 'Total Year to Maturity', i+2)},(100*{excel_ref(excel_df, 'kupon %', i+2)}),100,0)"
                        for i in range(len(excel_df))
                    ]
                    excel_df['Price if Yield Cut'] = [
                        f"=-PV(({excel_ref(excel_df, 'y mbi jual', i+2)}-{yield_cut_input/100.0}),{excel_ref(excel_df, 'Total Year to Maturity', i+2)},(100*{excel_ref(excel_df, 'kupon %', i+2)}),100,0)"
                        for i in range(len(excel_df))
                    ]
                
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    excel_df.to_excel(writer, index=False, sheet_name='Sheet1')
                    params_sheet = writer.book.create_sheet("_Parameters")
                    params_sheet["A1"] = "settlement_date"
                    params_sheet["B1"] = pd.to_datetime(settlement_date).date()
                    params_sheet.sheet_state = "hidden"
                
                st.download_button(
                    label="Unduh Data (Excel)",
                    data=buffer.getvalue(),
                    file_name="hasil_edit.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("Tabel tidak terdeteksi pada PDF ini.")
                
    except Exception as e:
        st.error(f"Gagal memproses file: {e}")
